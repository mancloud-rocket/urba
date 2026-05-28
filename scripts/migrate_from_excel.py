#!/usr/bin/env python3
"""Migra datos desde ctacts/ a seed.json para BALDE."""

import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

ROOT = Path(__file__).resolve().parents[1]
CTACTS = ROOT.parent / "ctacts"
OUT = ROOT / "server" / "data" / "seed.json"

ESTADO_VENTA_MAP = {
    "PAGO": "pagado",
    "DEBE": "debe",
    "DEBE ": "debe",
    "PEDIDO": "pedido",
    "USD 100": "parcial",
    "ENTREGADO": "entregado",
}

ESTADO_PAGO_MAP = {
    "PAGO": "pagado",
    "PENDIENTE": "pendiente",
    None: "pendiente",
}

MEDIO_MAP = {
    "EFECTIVO": "efectivo",
    "TRANSFERENCIA": "transferencia",
    "TARJETA": "tarjeta",
    "CHEQUE": "cheque",
}


def ser(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if v is None:
        return None
    return v


def norm_medio(v):
    if not v:
        return None
    key = str(v).strip().upper()
    return MEDIO_MAP.get(key, "otro")


def norm_estado_venta(v):
    if not v:
        return "otro"
    key = str(v).strip().upper()
    return ESTADO_VENTA_MAP.get(key, "otro")


def norm_estado_pago(v):
    if not v:
        return "pendiente"
    key = str(v).strip().upper()
    return ESTADO_PAGO_MAP.get(key, ESTADO_PAGO_MAP.get(v, "otro"))


def parse_iva(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().upper()
    if s == "PENDIENTE":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def migrate_facturas(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    wsd = wb["DATOS"]
    wsde = wb["DEUDORES"]

    clients = []
    for r in range(8, wsd.max_row + 1):
        cod = wsd.cell(r, 2).value
        if not cod:
            continue
        clients.append({
            "codigo": str(cod).strip(),
            "nombre": (wsd.cell(r, 5).value or "").strip(),
            "rut": ser(wsd.cell(r, 3).value),
            "identificacion": ser(wsd.cell(r, 4).value),
            "telefono": ser(wsd.cell(r, 6).value),
            "telefono2": ser(wsd.cell(r, 7).value),
            "direccion": wsd.cell(r, 8).value,
            "barrio": wsd.cell(r, 9).value,
            "email": wsd.cell(r, 10).value,
            "plazo_dias": int(wsd.cell(r, 12).value or 7),
        })

    entries = []
    for r in range(8, wsde.max_row + 1):
        cod = wsde.cell(r, 4).value
        if not cod:
            continue
        debe = wsde.cell(r, 7).value
        abono = wsde.cell(r, 8).value
        if abono:
            tipo = "abono"
            monto = float(abono)
        elif debe:
            tipo = "cargo"
            monto = float(debe)
        else:
            continue
        entries.append({
            "cliente_codigo": str(cod).strip(),
            "fecha": ser(wsde.cell(r, 2).value) or datetime.now().date().isoformat(),
            "referencia": ser(wsde.cell(r, 3).value),
            "tipo": tipo,
            "monto": monto,
            "medio_pago": norm_medio(wsde.cell(r, 9).value),
            "fecha_vencimiento": ser(wsde.cell(r, 10).value),
            "observacion": wsde.cell(r, 11).value,
            "estado": wsde.cell(r, 12).value,
        })

    return clients, entries


def migrate_electro(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    suppliers = list(wb.sheetnames)
    sales = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in range(6, ws.max_row + 1):
            cliente = ws.cell(r, 2).value
            producto = ws.cell(r, 4).value
            if not cliente or not producto:
                continue
            cod_prod = ws.cell(r, 3).value
            sales.append({
                "proveedor": sn,
                "cliente_nombre": str(cliente).strip(),
                "codigo_producto": str(cod_prod) if cod_prod is not None else None,
                "producto": str(producto).strip(),
                "usd_venta": float(ws.cell(r, 5).value or 0),
                "usd_costo": float(ws.cell(r, 7).value or 0),
                "iva": parse_iva(ws.cell(r, 9).value),
                "nro_factura": ser(ws.cell(r, 10).value),
                "modalidad": ws.cell(r, 12).value,
                "estado_venta": norm_estado_venta(ws.cell(r, 13).value),
                "fecha_venta": ser(ws.cell(r, 14).value),
                "estado_pago": norm_estado_pago(ws.cell(r, 21).value),
                "fecha_pago": ser(ws.cell(r, 23).value),
                "banco": ws.cell(r, 25).value,
                "nro_transaccion": ser(ws.cell(r, 26).value),
                "nro_recibo": ser(ws.cell(r, 22).value),
            })
    return suppliers, sales


def main():
    facturas = CTACTS / "FACTURAS PENDIENTES.xlsm"
    electro = CTACTS / "ELECTRODOMESTICOS.xlsx"
    if not facturas.exists():
        print(f"No encontrado: {facturas}")
        sys.exit(1)
    if not electro.exists():
        print(f"No encontrado: {electro}")
        sys.exit(1)

    clients, entries = migrate_facturas(facturas)
    suppliers, sales = migrate_electro(electro)

    # Vincular ventas a clientes por nombre aproximado
    by_name = {}
    for c in clients:
        key = re.sub(r"\s+", " ", c["nombre"].upper().strip())
        by_name[key] = c["codigo"]
        first = key.split()[0] if key else ""
        if first and first not in by_name:
            by_name[first] = c["codigo"]

    for s in sales:
        cn = re.sub(r"\s+", " ", s["cliente_nombre"].upper().strip())
        s["cliente_codigo"] = by_name.get(cn) or by_name.get(cn.split()[0] if cn else "", None)

    payload = {
        "migrated_at": datetime.now().isoformat(),
        "clients": clients,
        "ledger_entries": entries,
        "suppliers": [{"nombre": n} for n in suppliers],
        "sales_lines": sales,
        "allowed_phones": [
            {"telefono": "59899123456", "nombre": "Admin demo"},
        ],
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"Migracion OK -> {OUT}")
    print(f"  clientes: {len(clients)}")
    print(f"  movimientos: {len(entries)}")
    print(f"  proveedores: {len(suppliers)}")
    print(f"  ventas: {len(sales)}")
    unlinked = sum(1 for s in sales if not s.get("cliente_codigo"))
    if unlinked:
        print(f"  ventas sin cliente vinculado: {unlinked} (revisar nombres)")


if __name__ == "__main__":
    main()
